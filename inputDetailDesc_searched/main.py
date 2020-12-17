"""20200205_inputDetailDesc_inSuricataExcelFile"""
"""
엑셀파일에 상세설명을 작성하기 위한 코드
엑셀파일의 cve 코드에 대한 내용을 검색한 후 메모장에 작성
엑셀파일의 cve 코드와 메모장의 cve 코드 값을 비교한 후 같은 cve 코드 값일 경우
해당 cve의 설명을 엑셀파일의 상세 설명란에 작성
"""

"""1. 필요 모듈을 호출"""
# 정규식을 이용하기 위한 모듈
import re
# 파이썬으로 엑셀파일을 제어하기 위한 모듈
from openpyxl import load_workbook
# 에러의 내용이 작성된 딕셔너리를 저장하는 모듈
from rule_readMemo.inputDetailDesc_searched.inputErrorCVE import *

"""2. 메모장 관련"""
# 메모장 파일을 불러옴
memoFile = open("원하는 텍스트 파일 경로", encoding='UTF8')
# 메모장의 내용을 한 리스트에 삽입 > 메모장의 한 줄이 한 인덱스 안에 들어감
memoFile_linesList = memoFile.readlines()
# 메모장의 줄 수를 알아내기 위해서 리스트의 길이를 파악 > 전체 메모장 줄 수 -1 > 리스트는 인덱싱이 0부터 시작을 하기 때문
memoFile_totalCount = len(memoFile_linesList)
print(f'메모장 총 줄 수: {memoFile_totalCount}')
# 메모장 리스트의 인덱싱을 위한 변수 선언
memoFile_linesCount = 1
# 정규식을 활용하여 내용을 추출
# cve 코드를 불러오기 위한 정규식
regex_cveFind = re.compile(r"(.*?)ㅃ")
# 검색으로 찾은 cve 코드의 상세 설명을 위한 정규식
regex_detailDescFind = re.compile(r"ㅃ(.*?)ㅆ")

"""3. 엑셀파일 관련"""
# 엑셀파일을 불러옴
open_excelFile = load_workbook("원하는 엑셀 파일 경로")
# 해당 엑셀파일의 시트를 설정
load_excelSheet = open_excelFile['rule_info']
# 엑셀 파일의 전체 줄 수를 가져옴 > 반복문을 활용하기 위해서
rowCount_excelFile = load_excelSheet.max_row
print(f'엑셀 파일 줄 수: {rowCount_excelFile}')
# 반복문에 조건을 걸기 위해서 변수 선언
count_rowCount = 2

"""4. 혹시 에러가 발생했을 경우, 해당 내용을 저장하는 행"""
dic_different_cve = {}

"""5. 반복문"""
"""5-1. 메모장과 엑셀 파일의 줄 수가 같은지 먼저 확인"""
"""
먼저 엑셀 파일과 메모장 파일의 줄 수가 같다는 것을 표현
엑셀 파일의 cve 값을 그대로 가져왔기 때문에 줄 수가 같아야한다
그래야 엑셀 파일의 cve값과 메모장 파일의 cve 값을 1대1 매칭을 시킬 수가 있다. 
"""
if memoFile_totalCount == rowCount_excelFile:
    print('줄 수가 같다')
    """5-2. 반복문 실행 > 여기서는 엑셀 파일로 조건을 걸었음"""
    """
    엑셀 파일은 행 수가 1부터 시작을 하기 때문에
    조건을 걸 때, <로 조건을 걸 경우 맨 마지막 줄은 조건에 먹히지 않으므로
    <=를 사용해서 조건을 걸어야 한다.
    """
    while count_rowCount <= rowCount_excelFile:
        print(f'{count_rowCount}열: {load_excelSheet.cell(row=count_rowCount, column=6).value}')

        """5-3. cve값에 내용이 없을 경우"""
        """
        cve 값에 내용이 없을 경우에는 이미 상세 설명에 룰의 이름이 들어가 있기 때문에
        굳이 내용을 볼 필요가 없다
        게다가 해당 줄에는 빈 공백이기 때문에 더더욱 비교할 필요가 없다
        바로 줄 수를 올린다.
        """
        if load_excelSheet.cell(row=count_rowCount, column=6).value == None:
            print("넣을 값이 없다.")
            
        else:
            """5-4. cve 코드가 들어 있는 경우"""
            print(f'{count_rowCount}줄 내용: {memoFile_linesList[memoFile_linesCount]}')

            if load_excelSheet.cell(row=count_rowCount, column=9).value != None:
                print('이미 상세설명에 대한 값을 검색해서 넣었다')
            else:
                raw_cve_memoFile = regex_cveFind.search(memoFile_linesList[memoFile_linesCount])
                """
                여기서 try except 문을 사용한 이유
                아직 모든 cve 코드의 값을 검색한 것이 아니기 때문에
                검색하지 않은 부분에는 /가 들어갈 이유가 없다.
                그러므로 /를 찾는 정규식에는 에러가 발생할 수 밖에 없으며
                해당 에러가 발생한 경우에는 아래에 모든 cve 코드의 검색한 내용이 들어가 있지 않으므로
                전체 코드를 종료시킨다.
                """
                try:
                    """
                    cve 코드의 검색한 내용이 들어가 있을 경우
                    """
                    cve_memoFile = raw_cve_memoFile.group().replace('ㅃ','')

                    """5-5. 만일을 대비한 cve 코드의 값이 메모장과 엑셀 파일이 같은지 비교하는 조건문"""
                    """
                    현재 코드를 실행시키기에 먼저 필요한
                    메모장과 엑셀파일의 줄 수를 맞추기 위해서 먼저 메모장의 마지막 줄을 뺀 상황
                    그리고 엑셀파일의 cve 코드를 메모장에 옮기는 과정에서 오류가 발생했을 수도 있기에
                    작성한 조건문
                    """
                    if cve_memoFile == load_excelSheet.cell(row=count_rowCount, column=6).value:
                        print('cve 값이 같다')
                        """
                        cve 값이 같으면 해당 cve 코드의 검색한 내용을 엑셀파일에 정규식으로 추출한 뒤 삽입
                        """
                        raw_desc_memoFile = regex_detailDescFind.search(memoFile_linesList[memoFile_linesCount])
                        cve_descFile = raw_desc_memoFile.group().replace('ㅃ','').replace('ㅆ', '')

                        print(f'{count_rowCount}에 {cve_descFile} 넣는다.')

                        load_excelSheet.cell(row=count_rowCount, column=9).value = cve_descFile
                    else:
                        """
                        만일이라도 cve 코드의 값이 다른 경우 확인을 위해서 
                        새 메모장을 만들어서 해당 내용을 저장하고
                        엑셀파일을 저장함
                        """
                        print('cve 값이 다르다')
                        print(f'{count_rowCount}행 확인 바람')

                        dic_different_cve[count_rowCount] = load_excelSheet.cell(row=count_rowCount, column=6).value
                except AttributeError as attributeError:
                    print('아직 룰 검색을 안함')
                    print(attributeError)
                    break
                
        memoFile_linesCount = memoFile_linesCount + 1
        count_rowCount = count_rowCount + 1
else:
    print('줄 수가 다르다.')

"""6. 엑셀 파일을 저장"""
open_excelFile.save('저장할 엑셀 파일 이름')

"""7. 메모장 파일 저장"""
"""
엑셀파일을 만들면서
잘못된 부분을 딕셔너리에 저장하고
딕셔너리를 새 메모장에 저장함
"""
inputDifferentCVE(dic_different_cve)
