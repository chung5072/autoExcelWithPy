"""200123_multi_mainRunning.py"""
"""엑셀파일에 자동으로 등록해주는 코드"""

"""
먼저 메모장에 있는 값을 정규식을 활용하여 뽑아낸 다음
뽑아낸 값이 엑셈 파일의 sid 값과 같은 것이 있을 경우에는 
엑셀 파일의 특정 컬럼에 작성한다.
"""

"""1. 필요한 모듈 불러오기"""
"""1-1. 엑셀 파일을 제어"""
from openpyxl import load_workbook

"""1-2. 메모장에 있는 내용을 정규식으로 추출함"""
import re

"""1-3. 작업을 조금이나마 빨리 끝내기 위해서 멀티 프로세서를 사용"""
import multiprocessing
# 작업 시간이 얼마나 걸렸는지 확인
import time

"""2. 엑셀파일의 sid값과 메모장에 있는 sid 값을 비교하여 엑셀파일에 작성"""
def inputValue():
    # 작업 시간이 얼마나 걸렸는지 확인
    startTime = time.time() # 시작 시간

    # 엑셀 파일을 글로벌 변수로 선언하여 어느 위치에서든지 이용이 가능하도록
    global open_ruleFile

    """2-1. 메모장에 있는 값들을 추출하여 활용하기 위해서 선언"""
    # 먼저 해당 단어가 있는지 확인, 메모장을 읽다 보면 값이 없는 경우도 있기 때문에 필요
    word_sid = 'sid'
    word_cve = 'cve' # 사용 안함
    word_desc = 'desc' # 사용 안함
    # 정규식을 활용하여 원하는 값을 추출
    regex_sid_value = re.compile(r'"sid": "(.*?)"')
    regex_cve_value = re.compile(r'"cve": "(.*?)"')
    regex_desc_value = re.compile(r'"desc": "(.*?)"')
    # find 함수를 활용하여 값이 있는지 확인, find함수로 원하는 문구의 위치를 알 수 있음
    find_word_sid = None
    find_word_cve = None
    find_word_desc = None

    """2-2. 메모장을 염"""
    memoFile = open("원하는 텍스트 파일 경로") # 경로는 전체 경로를 사용
    # 메모장을 열어서 메모장을 읽기 위한 변수를 설정
    memoFile_linesList = memoFile.readlines() # 메모장의 내용을 한 리스트에 전부 넣음
    memoFile_totalCount = len(memoFile_linesList) # 전체 메모장의 줄 수를 알 수 있음
    print(memoFile_totalCount)
    memoFile_linesCount = 0 # 조건문을 활용하기 위해서 선언한 변수

    """2-3. 엑셀 파일을 염"""
    # 엑셀파일을 로드함.
    open_ruleFile = load_workbook("원하는 엑셀 파일 경로") # 경로는 전체 경로를 사용
    load_ruleSheet = open_ruleFile['rule_info'] # 엑셀파일의 특정 이름을 가진 시트를 염
    # 엑셀 파일의 전체 줄 수를 파악
    rowCount_ruleFile = load_ruleSheet.max_row
    print(rowCount_ruleFile)
    count_rowCount = 1 # 조건문을 활용하기 위해서 선언한 변수

    """2-4. 조건문을 활용하여 메모장과 엑셀파일의 sid 값을 비교"""
    while count_rowCount < rowCount_ruleFile: # 엑셀 파일 줄 수
        while memoFile_linesCount < memoFile_totalCount: # 메모장 파일 줄 수
            """2-5. sid값이 해당 줄에 있는지 먼저 확인."""
            """
            사실 이번 메모장에서는 굳이 필요한 코드는 아니었음.
            빽빽한 줄에 전부 sid라는 문자열이 들어있었기 때문에.
            """
            find_word_sid = memoFile_linesList[memoFile_linesCount].find(word_sid)
            # 엑셀 파일의 컬럼에 접근하여 원하는 값을 빼냄
            excel_sid = load_ruleSheet.cell(row=count_rowCount + 1, column=7).value
            # sid 값이 어떤 것은 문자열 타입이고 어떤 것은 int 타입이었기 때문에 하나로 맞춰줌
            if type(excel_sid) == int: # 공백이 없이 전부 숫자만 있을 때는 int로 나왔음
                str_excelSid = str(excel_sid)
            elif type(excel_sid) == str:
                str_excelSid = excel_sid.replace(' ', '') # 공백이 섞여 있었으면 문자열로 나왔음
            excel_cve = load_ruleSheet.cell(row=count_rowCount + 1, column=6).value
            str_excelCve = str(excel_cve)

            if find_word_sid > -1:
                raw_matchType_regex_sid_value = regex_sid_value.search(memoFile_linesList[memoFile_linesCount])
                raw_regex_sid_value = raw_matchType_regex_sid_value.group() # 정규식의 search를 사용하고 group을 꼭 해줘야 값이 잘 나옴
                result_regex_sid_value = raw_regex_sid_value.replace('"sid": "', '').replace('"', '')
                print(result_regex_sid_value)

                """2-6. sid값이 같을 경우에는 cve 값과 desc값을 넣음"""
                if str_excelSid == result_regex_sid_value:
                    raw_matchType_regex_cve_value = regex_cve_value.search(memoFile_linesList[memoFile_linesCount])
                    raw_regex_cve_value = raw_matchType_regex_cve_value.group()
                    result_regex_cve_value = raw_regex_cve_value.replace('"cve": "', '').replace('"', '')

                    if str_excelCve == result_regex_cve_value:
                        print('same cve')
                        pass
                    else:
                        print(f'input cve: {result_regex_cve_value}')
                        """2-7. 값을 원하는 컬럼에 입력"""
                        load_ruleSheet.cell(row=count_rowCount + 1, column=6).value = result_regex_cve_value

                    raw_matchType_regex_desc_value = regex_desc_value.search(memoFile_linesList[memoFile_linesCount])
                    raw_regex_desc_value = raw_matchType_regex_desc_value.group()
                    result_regex_desc_value = raw_regex_desc_value.replace('"desc": "', '').replace('"', '')
                    print(f'input desc: {result_regex_desc_value}')

                    """2-7. 값을 원하는 컬럼에 입력"""
                    load_ruleSheet.cell(row=count_rowCount + 1, column=12).value = result_regex_desc_value

                else:
                    print(f'{str(excel_sid)}와 {result_regex_sid_value}는 다름')
                    pass
            else:
                pass

            memoFile_linesCount = memoFile_linesCount + 1

        memoFile_linesCount = 0 # 만일 0으로 초기화를 안하고 돌리면 전부 비교가 불가능하다. 다 더해진 상태로 다음 조건문을 돌기 때문에
        count_rowCount = count_rowCount + 1


    endTime = time.time() # 코드 실행시간을 체크
    print(endTime - startTime)

    """2-8. 모든 내용을 입력하고 나서 엑셀파일을 저장."""
    open_ruleFile.save("저장할 엑셀 파일 경로")

if __name__ == '__main__':
    """3. 멀티 프로세서를 활용하여 함수 실행."""
    p = multiprocessing.Process(target=inputValue)
    p.start()









