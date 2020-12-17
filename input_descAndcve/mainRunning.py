import re
import time

from openpyxl import load_workbook
startTime = time.time()

word_sid = 'sid'
word_cve = 'cve'
word_desc = 'desc'

regex_sid_value = re.compile(r'"sid": "(.*?)"')
regex_cve_value = re.compile(r'"cve": "(.*?)"')
regex_desc_value = re.compile(r'"desc": "(.*?)"')

find_word_sid = None
find_word_cve = None
find_word_desc = None

memoFile = open("원하는 텍스트 파일 경로")

memoFile_linesList = memoFile.readlines()
memoFile_totalCount = len(memoFile_linesList)
print(memoFile_totalCount)
memoFile_linesCount = 0

open_ruleFile = load_workbook("원하는 엑셀파일 경로")
load_ruleSheet = open_ruleFile['rule_info']

rowCount_ruleFile = load_ruleSheet.max_row
print(rowCount_ruleFile)
count_rowCount = 1

while count_rowCount < rowCount_ruleFile:
    while memoFile_linesCount < memoFile_totalCount:
        find_word_sid = memoFile_linesList[memoFile_linesCount].find(word_sid)
        excel_sid = load_ruleSheet.cell(row=count_rowCount + 1, column=7).value
        if type(excel_sid) == int:
            str_excelSid = str(excel_sid)
        elif type(excel_sid) == str:
            str_excelSid = excel_sid.replace(' ', '')
        excel_cve = load_ruleSheet.cell(row=count_rowCount + 1, column=6).value
        str_excelCve = str(excel_cve)

        if find_word_sid > -1:
            raw_matchType_regex_sid_value = regex_sid_value.search(memoFile_linesList[memoFile_linesCount])
            raw_regex_sid_value = raw_matchType_regex_sid_value.group()
            result_regex_sid_value = raw_regex_sid_value.replace('"sid": "', '').replace('"', '')
            print(result_regex_sid_value)

            if str_excelSid == result_regex_sid_value:
                raw_matchType_regex_cve_value = regex_cve_value.search(memoFile_linesList[memoFile_linesCount])
                raw_regex_cve_value = raw_matchType_regex_cve_value.group()
                result_regex_cve_value = raw_regex_cve_value.replace('"cve": "', '').replace('"', '')

                if str_excelCve == result_regex_cve_value:
                    print('same cve')
                    pass
                else:
                    print(f'input cve: {result_regex_cve_value}')
                    load_ruleSheet.cell(row=count_rowCount + 1, column=6).value = result_regex_cve_value

                raw_matchType_regex_desc_value = regex_desc_value.search(memoFile_linesList[memoFile_linesCount])
                raw_regex_desc_value = raw_matchType_regex_desc_value.group()
                result_regex_desc_value = raw_regex_desc_value.replace('"desc": "', '').replace('"', '')
                print(f'input desc: {result_regex_desc_value}')
                load_ruleSheet.cell(row=count_rowCount + 1, column=12).value = result_regex_desc_value

            else:
                print(f'{str(excel_sid)}와 {result_regex_sid_value}는 다름')
                pass
        else:
            pass

        memoFile_linesCount = memoFile_linesCount + 1

    memoFile_linesCount = 0
    count_rowCount = count_rowCount + 1

endTime = time.time()
print(endTime - startTime)

open_ruleFile.save("저장하고자 하는 엑셀 파일 경로")







