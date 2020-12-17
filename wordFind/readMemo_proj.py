import re
import openpyxl
from openpyxl import load_workbook

from rule_readMemo.wordFind.regex_extraFind import *
from rule_readMemo.wordFind.regex_msgFind import *
from rule_readMemo.wordFind.regex_referenceFind import *
from rule_readMemo.wordFind.regex_classtypeFind import *
from rule_readMemo.wordFind.regex_sidFind import *
from rule_readMemo.wordFind.regex_metadataFind import *
from rule_readMemo.wordFind.regex_cveFind import *
from rule_readMemo.wordFind.regex_referenceDocFind import *
from rule_readMemo.wordFind.regex_revFind import *

word_extraFind = 'alert'
word_msgFind = 'msg:'
word_referenceFind = 'reference:url,w'
word_cveFind = 'reference:cve'
word_classtypeFind = 'classtype:'
word_sidFind = 'sid:'
word_metadataFind = 'metadata:'
# 내용을 빼기 위해서 필요한 단어 찾기
word_referenceDocFind = 'reference:url,doc'
word_version = 'rev:'
word_lastSentence = ' )'

# 내용을 넣기 위해서 필요한 정규식
regex_extraFind_Line = re.compile(r".+?\(")
regex_msgFind_Line = re.compile(r'msg:(.*?); ')
regex_referenceFind_Line = re.compile(r'reference:url,w(.*?); ')
regex_cveFind_Line = re.compile(r'reference:cve,(.*?); ')
regex_classtypeFind_Line = re.compile(r'classtype:(.*?); ')
regex_sidFind_Line = re.compile(r'sid:(.*?); ')
regex_metadataFind_Line = re.compile(r'metadata:(.*?);')
# 내용을 빼기 위해서 필요한 정규식
regex_referenceDocFind = re.compile(r'reference:url,doc(.*?); ')
regex_version = re.compile(r'rev:(.*?); ')

memoFile = open("원하는 텍스트파일 경로")
continue_read_memoFile = None

memoFile_linesList = memoFile.readlines()
memoFile_totalCount = len(memoFile_linesList)
memoFile_linesCount = 0
excelFile_linesCount = 1

writeSheet = openpyxl.Workbook()

testSheet = writeSheet.active
testSheet.title = 'rule_info'

testSheet.cell(row=1, column=1).value = 'Rule Name'
testSheet.cell(row=1, column=2).value = 'Rule Setting info'
testSheet.cell(row=1, column=3).value = 'classification'
testSheet.cell(row=1, column=4).value = 'Risk'
testSheet.cell(row=1, column=5).value = 'CVE'
testSheet.cell(row=1, column=6).value = 'SID'
testSheet.cell(row=1, column=7).value = '생성정보'
testSheet.cell(row=1, column=8).value = '상세 설명'
testSheet.cell(row=1, column=9).value = 'Reference'

while memoFile_linesCount < memoFile_totalCount:
    if memoFile_linesList[memoFile_linesCount] != '\n':
        print(f"{memoFile_linesCount+1}번째 줄 \n{memoFile_linesList[memoFile_linesCount]}")

        hitWord_msgFind = memoFile_linesList[memoFile_linesCount].find(word_msgFind)
        if hitWord_msgFind > -1:
            print('Rule Name')
            result_msgFind = msg(memoFile_linesList[memoFile_linesCount], regex_msgFind_Line)
            result_msgFind_extractContents = result_msgFind.replace('msg:"', '').replace('";', '')
            print(result_msgFind_extractContents)
            testSheet.cell(row=excelFile_linesCount, column=1).value = result_msgFind_extractContents

        hitWord_referenceFind = memoFile_linesList[memoFile_linesCount].find(word_referenceFind)
        if hitWord_referenceFind > -1:
            print('Reference')
            result_referenceFind = reference(memoFile_linesList[memoFile_linesCount], regex_referenceFind_Line)
            result_referenceFind_extractContents = result_referenceFind.replace('reference:url,', '').replace(';', '')
            print(result_referenceFind_extractContents)
            testSheet.cell(row=excelFile_linesCount, column=9).value = result_referenceFind_extractContents

        hitWord_cveFind = memoFile_linesList[memoFile_linesCount].find(word_cveFind)
        if hitWord_cveFind > -1:
            print('CVE')
            result_cveFind = cve(memoFile_linesList[memoFile_linesCount], regex_cveFind_Line)
            result_cveFind_extractContents = result_cveFind.replace('reference:cve,', '').replace(';', '')
            print(result_cveFind_extractContents)
            testSheet.cell(row=excelFile_linesCount, column=5).value = result_cveFind_extractContents

        hitWord_classtypeFind = memoFile_linesList[memoFile_linesCount].find(word_classtypeFind)
        if hitWord_classtypeFind > -1:
            print('classification')
            result_classtypeFind = classtype(memoFile_linesList[memoFile_linesCount], regex_classtypeFind_Line)
            result_classtypeFind_extractContents = result_classtypeFind.replace('classtype:','').replace(';', '')
            print(result_classtypeFind_extractContents)
            testSheet.cell(row=excelFile_linesCount, column=3).value = result_classtypeFind_extractContents

        hitWord_sidFind = memoFile_linesList[memoFile_linesCount].find(word_sidFind)
        if hitWord_sidFind > -1:
            print('sid')
            result_sidFind = sid(memoFile_linesList[memoFile_linesCount], regex_sidFind_Line)
            result_sidFind_extractContents = result_sidFind.replace('sid:', '').replace(';', '')
            print(result_sidFind_extractContents)
            testSheet.cell(row=excelFile_linesCount, column=6).value = result_sidFind_extractContents

        hitWord_metadataFind = memoFile_linesList[memoFile_linesCount].find(word_metadataFind)
        if (hitWord_metadataFind > -1):
            print('생성 정보_전방 탐색')
            result_metadataFind = metadata(memoFile_linesList[memoFile_linesCount], regex_metadataFind_Line)
            result_metadataFind_extractContents = result_metadataFind.replace('metadata:', '').replace(';', '')
            print(result_metadataFind_extractContents)
            testSheet.cell(row=excelFile_linesCount, column=7).value = result_metadataFind_extractContents

        hitWord_referenceDocFind = memoFile_linesList[memoFile_linesCount].find(word_referenceDocFind)
        if hitWord_referenceDocFind > -1:
            print('빼기 위한 내용 - reference doc')
            result_referenceDocFind = referenceDoc(memoFile_linesList[memoFile_linesCount], regex_referenceDocFind)
            print(result_referenceDocFind)

        hitWord_revFind = memoFile_linesList[memoFile_linesCount].find(word_version)
        if hitWord_revFind > -1:
            print('빼기 위한 내용 - rev')
            result_revFind = rev(memoFile_linesList[memoFile_linesCount], regex_version)
            print(result_revFind)

        hitWord_extraFind = memoFile_linesList[memoFile_linesCount].find(word_extraFind)
        if hitWord_extraFind > -1:
            print('Rule Setting Info')
            result_extraFind = extra(memoFile_linesList[memoFile_linesCount], regex_extraFind_Line)
            # print(result_extraFind)

            try:
                result_ruleSettingInfo = memoFile_linesList[memoFile_linesCount].replace(result_extraFind, '').replace(
                    result_msgFind, '').replace(result_referenceFind, '').replace(result_referenceDocFind, '').replace(
                    result_cveFind, '').replace(result_classtypeFind, '').replace(result_sidFind, '').replace(
                    result_revFind, '').replace(result_metadataFind, '').replace(word_lastSentence, '')
            except NameError:
                result_ruleSettingInfo = memoFile_linesList[memoFile_linesCount].replace(result_extraFind, '').replace(
                    result_msgFind, '').replace(result_referenceFind, '').replace(result_referenceDocFind, '').replace(
                    result_classtypeFind, '').replace(result_sidFind, '').replace(result_revFind, '').replace(
                    result_metadataFind, '').replace(word_lastSentence, '')
            print(result_ruleSettingInfo)
            testSheet.cell(row=excelFile_linesCount, column=2).value = result_ruleSettingInfo

        excelFile_linesCount = excelFile_linesCount + 1
    elif memoFile_linesList[memoFile_linesCount] == '\n':
        pass
        # choose_continue = input('continue? ')
        # if choose_continue == 'y':
        #     pass
        # elif choose_continue == 'n':
        #     # writeSheet.save(filename='test_ruleFile_write3.xlsx')
        #     break

    memoFile_linesCount = memoFile_linesCount + 1


writeSheet.save(filename="원하는 엑셀파일 경로")