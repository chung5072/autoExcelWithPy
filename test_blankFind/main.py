"""200207_findBlankField_inSuricataRuleExcelFile_inDetailDesc"""
"""
특정 내용을 추출하길 원하는 엑셀파일에서 상세설명에 들어간 내용이 공백인 칸이 어디인지
해당 행과 해당 cve 코드의 내용을 받아서 출력하는 코드
"""

"""1. 필요한 모듈을 불러옴"""
# 파이썬으로 엑셀파일을 제어하기 위한 모듈
from openpyxl import load_workbook

"""2. 엑셀파일 관련"""
# 제어하기 원하는 엑셀파일을 불러옴
open_ruleFile = load_workbook("원하는 엑셀 파일 경로")
# 해당 엑셀파일의 시트를 설정함
load_ruleSheet = open_ruleFile['rule_info']
# 엑셀 파일의 전체를 보기 위해서 해당 엑셀파일의 전체 행 수를 가져옴
rowCount_ruleFile = load_ruleSheet.max_row
# 반복문을 활용하기 위해서 설정한 변수 > 가장 윗 줄은 제목이 들어간 곳이니 제외
count_rowCount = 2

"""3. 빈 칸인 행을 넣기 위해서 선언한 딕셔너리"""
blank_detailDesc_dict = {}

"""4. 반복문"""
while count_rowCount <= rowCount_ruleFile:
    """4-1. 특정 컬럼을 가져오는 변수"""
    getDetailDesc = load_ruleSheet.cell(row=count_rowCount, column=9).value
    getDetailCVE = load_ruleSheet.cell(row=count_rowCount, column=6).value

    """4-2. 현재 어디까지 확인했는지 알려주는 출력문"""
    print(f'{count_rowCount}행 확인중')

    """4-3. 해당 엑셀파일의 값이 들어가 있지 않다는 것을 알려주는 조건문"""
    if getDetailDesc == None:
        """4-4. 엑셀파일에 값이 들어가 있지 않으면 해당 행과 cve 코드의 값을 딕셔너리에 저장함"""
        blank_detailDesc_dict[count_rowCount] = getDetailCVE

    count_rowCount = count_rowCount + 1

"""5. 저장한 딕셔너리의 값을 출력"""
print(str(blank_detailDesc_dict))
        
    
