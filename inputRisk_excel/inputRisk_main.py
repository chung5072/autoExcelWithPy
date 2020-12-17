"""200129_inputRisk_inSuricataRule"""
"""
특정 내용을 추출하길 원하는 엑셀파일의 classtype 값을 가져와서 
해당 classtype에 매칭되는 risk의 수치를 가져와서
1이면 높음, 2면 중간, 3이면 낮음을 엑셀 파일에 저장한다.
"""

"""1. 필요 모듈을 호출"""
# 엑셀 파일을 제어하기 위한 모듈을 호출
from openpyxl import load_workbook
# 리스크의 내용을 넣기 위해서 작성한 코드를 호출
from rule_readMemo.inputRisk_excel.inputRisk_memoForRisk import *

"""2. 엑셀 파일 관련"""
# 엑셀 파일을 불러옴
open_ruleFile = load_workbook("원하는 엑셀 파일 경로")
# 해당 엑셀 파일의 시트를 지정함
load_ruleSheet = open_ruleFile['rule_info']
# 엑셀파일을 전체 줄 수를 가져옴
rowCount_ruleFile = load_ruleSheet.max_row
# 반복문을 제어하기 위해서 선언한 변수
count_rowCount = 1

"""2. 반복문 실행"""
while count_rowCount < rowCount_ruleFile:
    """2-1. 엑셀파일의 classtype을 가져옴"""
    getClassification = load_ruleSheet.cell(row=count_rowCount+1, column=3).value
    print(f'{count_rowCount+1}줄의 위험도 측정: {getClassification}')

    """2-2. 해당 classtype을 활용하여 risk의 값을 가져오는 함수를 실행"""
    getRisk = memoForRisk(getClassification)
    print(getRisk)

    """2-3. 가져온 risk 값을 엑셀파일에 삽입"""
    load_ruleSheet.cell(row=count_rowCount+1, column=4).value = getRisk

    count_rowCount = count_rowCount + 1

"""3. 엑셀파일을 저장"""
open_ruleFile.save("저장할 엑셀 파일 이름")
